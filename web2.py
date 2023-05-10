import streamlit as st
import pandas as pd
from openpyxl import load_workbook

angka = 0
def main():
    st.header("WELCOME TO CHANIAGO GRUP")
    st.title("masukan inputan :")
    angka = st.number_input("masukan jumlah karung", 0, 100)
    st.button("masukan")
while True:
    if st.button == True:
     data = {"angka":[angka]}
     datafile = pd.DataFrame(data, columns=["angka"])
     datafile.to_excel("data_htg.xlsx", index=False)
     filename ='data_htg.xlsx'
     try:
         buku = load_workbook(filename)
         writer = pd.ExcelWriter(filename,engine='openpyxl')
         writer.book = book
         writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

         data.to_excel(writer, index = False, header = False,startrow = writer.sheet['sheet1'].max_row)

         writer.save()
         writer.close()
         st.success('berhasil')
     except FileNotFoundError:
         data.to-excel("baru.xlsx",index = False)
         st.success('file baru')
     st.success("berhasil")
    elif angka == 0:
        break
    st.success("telah masuk")
main()